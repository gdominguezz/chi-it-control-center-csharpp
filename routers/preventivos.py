from fastapi import APIRouter, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
from pathlib import Path
from fastapi.responses import HTMLResponse
from dateutil.relativedelta import relativedelta
import hashlib, json, io
import pandas as pd
import qrcode
from database import get_connection
from urllib.parse import quote
from PIL import Image, ImageDraw, ImageFont
from openpyxl.styles import PatternFill
from fastapi import Request
import zipfile
import io
# ───────── CARPETA QR ─────────
QR_DIR = Path("QR_CODES/MESAS")
QR_DIR.mkdir(parents=True, exist_ok=True)


router = APIRouter()

# ── Constantes ───────────────────────────────────────────
PDF_DIR = Path("PDF_DATABASE/PREVENTIVOS")
PDF_DIR.mkdir(parents=True, exist_ok=True)


# ── Modelo ───────────────────────────────────────────────
class Preventivo(BaseModel):
    ID_EQUIPO:         Optional[str] = None
    UBICACION:         Optional[str] = None
    PLAZO:             Optional[str] = None
    REALIZADO_POR:     Optional[str] = None
    FECHA_REALIZACION: Optional[str] = None
    OBSERVACIONES:     Optional[str] = None
    nombre_dispositivo:Optional[str] = None
    PLANTA:            Optional[str] = None
    CATEGORIA_COLOR:   Optional[str] = None
    ANIO_CREACION:     Optional[int] = None

# ── Tabla de auditoría (se crea al importar el router) ───
def _crear_tabla_auditoria():
    from database import db_conn
    try:
        with db_conn() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS public.auditoria_preventivos (
                        id                SERIAL PRIMARY KEY,
                        registro_id       INTEGER NOT NULL,
                        fecha_cambio      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        usuario           VARCHAR(100),
                        registro_anterior TEXT,
                        registro_nuevo    TEXT
                    )
                """)
                conn.commit()
            finally:
                cursor.close()
    except Exception as e:
        print(f"Error creando AUDITORIA_PREVENTIVOS: {e}")

_crear_tabla_auditoria()


# ── Context manager de conexión segura ──────────────────
# Usa db_conn() de database.py que llama putconn() correctamente
from database import db_conn as _db


# ── Helpers ──────────────────────────────────────────────
def _registrar_auditoria(registro_id, usuario, anterior, nuevo):
    try:
        with _db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO public.auditoria_preventivos
                (registro_id,usuario,registro_anterior,registro_nuevo,fecha_cambio)
                VALUES (%s,%s,%s,%s,%s)
            """, (registro_id, usuario,
                  json.dumps(anterior, default=str),
                  json.dumps(nuevo,    default=str),
                  datetime.now()))
            conn.commit()
            cursor.close()
    except Exception as e:
        print(f"Error auditoria preventivo: {e}")


def _obtener_usuario(request: Request, usuario_query: str = None) -> str:
    """Obtiene el usuario desde cookie, header X-Usuario o query param."""
    try:
        usr = request.cookies.get("usuario")
        if usr: return usr
        usr = request.headers.get("X-Usuario")
        if usr: return usr
    except Exception:
        pass
    return usuario_query or "SISTEMA"


def _pdf_path(id: int) -> Path:
    return PDF_DIR / f"{id}.pdf"



@router.get("/PREVENTIVOS")
def obtener_preventivos(
    page: int = Query(1),
    limit: int = Query(10),
    ID_EQUIPO: Optional[str] = Query(None),
    UBICACION: Optional[str] = Query(None),
    nombre_dispositivo: Optional[str] = Query(None),
    PLANTA: Optional[str] = Query(None),
    CATEGORIA_COLOR: Optional[str] = Query(None),
    ANIO_CREACION: Optional[str] = Query(None)
):

    where = "WHERE 1=1"
    params = []

    if ID_EQUIPO:
        where += ' AND id_equipo ILIKE %s'
        params.append(f"%{ID_EQUIPO}%")

    if UBICACION:
        where += ' AND ubicacion ILIKE %s'
        params.append(f"%{UBICACION}%")

    if nombre_dispositivo:
        where += ' AND nombre_dispositivo ILIKE %s'
        params.append(f"%{nombre_dispositivo}%")

    if PLANTA:
        where += ' AND planta ILIKE %s'
        params.append(f"%{PLANTA}%")

    if CATEGORIA_COLOR:
        where += ' AND categoria_color ILIKE %s'
        params.append(f"%{CATEGORIA_COLOR}%")

    if ANIO_CREACION and ANIO_CREACION.strip():
        try:
            where += ' AND anio_creacion = %s'
            params.append(int(ANIO_CREACION))
        except ValueError:
            pass

    with _db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(
                f'SELECT COUNT(*) FROM public.mantenimientos_preventivos {where}',
                params
            )
            total = cursor.fetchone()[0]

            offset = (page - 1) * limit

            cursor.execute(f"""
                SELECT *,
                CASE WHEN pdf IS NOT NULL THEN true ELSE false END AS tiene_pdf
                FROM public.mantenimientos_preventivos
                {where}
                ORDER BY
                  CASE LOWER(categoria_color)
                    WHEN 'verde'    THEN 1
                    WHEN 'gris'     THEN 2
                    WHEN 'azul'     THEN 3
                    WHEN 'rojo'     THEN 4
                    WHEN 'amarillo' THEN 5
                    WHEN 'rosa'     THEN 6
                    ELSE 7
                  END,
                  id DESC
                LIMIT %s OFFSET %s
            """, params + [limit, offset])

            columnas = [d[0] for d in cursor.description]
            data = [dict(zip(columnas, row)) for row in cursor.fetchall()]
        finally:
            cursor.close()

    return {
        "data": data,
        "total": total,
        "page": page
    }
# ════════════════════════════════════════════════════════
# JALAR LOS DATOS DEL PREVENTIVO 
# ════════════════════════════════════════════════════════
@router.get("/PREVENTIVO/DATOS/{id}")
def obtener_datos_preventivo(id: int, usuario: str = None):

    with _db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
            SELECT ubicacion, planta
            FROM public.mantenimientos_preventivos
            WHERE id=%s
            """, (id,))

            base = cursor.fetchone()
            if not base:
                return {"error": "Registro no encontrado"}

            ubicacion = base[0]
            planta    = base[1]

            cursor.execute("""
            SELECT id, id_equipo, nombre_dispositivo
            FROM public.mantenimientos_preventivos
            WHERE ubicacion=%s
            """, (ubicacion,))

            rows = cursor.fetchall()

            pc=""; impresora=""; ups=""; portatil=""
            id_pc=None; id_portatil=None; id_impresora=None; id_ups=None

            for r in rows:
                reg_id      = r[0]
                equipo      = r[1]
                dispositivo = (r[2] or "").upper()

                # Portátil/Laptop → columna "Portátil" del formato
                if "PORTATIL" in dispositivo or "LAPTOP" in dispositivo:
                    portatil    = equipo
                    id_portatil = reg_id

                # Computadora de escritorio / CPU → columna "PC" del formato
                elif "COMPUTADORA" in dispositivo or "CPU" in dispositivo:
                    pc    = equipo
                    id_pc = reg_id

                if "IMPRESORA" in dispositivo:
                    impresora    = equipo
                    id_impresora = reg_id

                if "UPS" in dispositivo:
                    ups    = equipo
                    id_ups = reg_id

            nombre_usuario = ""
            if usuario:
                cursor.execute("""
                SELECT nombre FROM public.usuarios WHERE usuario=%s
                """, (usuario.upper(),))
                user_row = cursor.fetchone()
                if user_row:
                    nombre_usuario = user_row[0]
        finally:
            cursor.close()

    return {
        "planta":       planta,
        "ubicacion":    ubicacion,
        "usuario":      nombre_usuario,
        "portatil":     portatil,
        "pc":           pc,
        "impresora":    impresora,
        "ups":          ups,
        "id_portatil":  id_portatil,
        "id_pc":        id_pc,
        "id_impresora": id_impresora,
        "id_ups":       id_ups
    }
# ════════════════════════════════════════════════════════
# GET  /PREVENTIVOS/{id}/HISTORIAL
# ════════════════════════════════════════════════════════
@router.get("/PREVENTIVOS/{id}/HISTORIAL")
def obtener_historial_preventivo(id: int):
    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    SELECT id,id_equipo,ubicacion,plazo,realizado_por,
                           fecha_realizacion,observaciones,nombre_dispositivo,planta,categoria_color,anio_creacion
                    FROM public.mantenimientos_preventivos WHERE id=%s
                """, (id,))
                row = cursor.fetchone()
                if not row:
                    return {"success": False, "error": "Registro no encontrado"}

                cols_actual = [d[0] for d in cursor.description]
                reg_actual  = dict(zip(cols_actual, row))

                cursor.execute("""
                    SELECT id,fecha_cambio,usuario,registro_anterior,registro_nuevo
                    FROM public.auditoria_preventivos
                    WHERE registro_id=%s ORDER BY fecha_cambio DESC
                """, (id,))

                historial = []
                for r in cursor.fetchall():
                    historial.append({
                        "id":                r[0],
                        "fecha":             r[1].isoformat() if r[1] else None,
                        "usuario":           r[2],
                        "registro_anterior": json.loads(r[3]) if r[3] else {},
                        "registro_nuevo":    json.loads(r[4]) if r[4] else {}
                    })
            finally:
                cursor.close()

        return {"success": True, "registro_actual": reg_actual, "historial": historial}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ════════════════════════════════════════════════════════
# POST /PREVENTIVO
# ════════════════════════════════════════════════════════
@router.post("/PREVENTIVO")
def crear_preventivo(data: Preventivo):
    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    INSERT INTO public.mantenimientos_preventivos
                    (id_equipo,ubicacion,plazo,realizado_por,
                     fecha_realizacion,observaciones,nombre_dispositivo,planta,categoria_color,anio_creacion)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                """, (data.ID_EQUIPO, data.UBICACION, data.PLAZO, data.REALIZADO_POR,
                      data.FECHA_REALIZACION, data.OBSERVACIONES,
                      data.nombre_dispositivo, data.PLANTA, data.CATEGORIA_COLOR, data.ANIO_CREACION))
                new_id = cursor.fetchone()[0]
                conn.commit()
            finally:
                cursor.close()
        return {"id": new_id}
    except Exception as e:
        return {"error": str(e)}
    


@router.get("/QR_GENERAR_TODOS")
def generar_qr_todos():

    base = "http://172.24.104.166:8000"

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w") as zipf:

        with _db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT DISTINCT ubicacion
                FROM public.mantenimientos_preventivos
                WHERE ubicacion IS NOT NULL
            """)
            ubicaciones = cursor.fetchall()
            cursor.close()

        for u in ubicaciones:
            ubicacion = u[0]

            try:
                ubicacion_url = quote(ubicacion)
                url = f"{base}/preventivos/qr/{ubicacion_url}"

                qr = qrcode.make(url).convert("RGB")

                qr_width, qr_height = qr.size
                altura_total = qr_height + 60

                img = Image.new("RGB", (qr_width, altura_total), "white")
                img.paste(qr, (0, 0))

                draw = ImageDraw.Draw(img)

                try:
                    font = ImageFont.truetype("arial.ttf", 22)
                except:
                    font = ImageFont.load_default()

                text_width = draw.textlength(ubicacion, font=font)
                x = (qr_width - text_width) / 2
                y = qr_height + 15

                draw.text((x, y), ubicacion, fill="black", font=font)

                img_buffer = io.BytesIO()
                img.save(img_buffer, format="PNG")

                zipf.writestr(f"{ubicacion}.png", img_buffer.getvalue())

            except Exception as e:
                print("Error QR:", ubicacion, e)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/x-zip-compressed",
        headers={
            "Content-Disposition": "attachment; filename=QRS_PREVENTIVOS.zip"
        }
    )

# ════════════════════════════════════════════════════════
# PUT  /PREVENTIVO/{id}
# ════════════════════════════════════════════════════════
@router.put("/PREVENTIVO/{id}")
def editar_preventivo(
    id: int, data: Preventivo, request: Request,
    usuario: Optional[str] = Query(None)
):
    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    SELECT id_equipo,ubicacion,plazo,realizado_por,
                           fecha_realizacion,observaciones,nombre_dispositivo,planta,categoria_color,anio_creacion
                    FROM public.mantenimientos_preventivos WHERE id=%s
                """, (id,))
                row = cursor.fetchone()
                anterior = {
                    "id_equipo": row[0], "ubicacion": row[1], "plazo": row[2],
                    "realizado_por": row[3],
                    "fecha_realizacion": str(row[4]) if row[4] else None,
                    "observaciones": row[5], "nombre_dispositivo": row[6],
                    "planta": row[7], "categoria_color": row[8],
                    "anio_creacion": row[9]
                } if row else {}

                cursor.execute("""
                    UPDATE public.mantenimientos_preventivos SET
                    id_equipo=%s,ubicacion=%s,plazo=%s,realizado_por=%s,
                    fecha_realizacion=%s,observaciones=%s,
                    nombre_dispositivo=%s,planta=%s,categoria_color=%s,anio_creacion=%s
                    WHERE id=%s
                """, (data.ID_EQUIPO, data.UBICACION, data.PLAZO, data.REALIZADO_POR,
                      data.FECHA_REALIZACION, data.OBSERVACIONES,
                      data.nombre_dispositivo, data.PLANTA, data.CATEGORIA_COLOR, data.ANIO_CREACION, id))
                conn.commit()
            finally:
                cursor.close()

        nuevo = {
            "id_equipo": data.ID_EQUIPO, "ubicacion": data.UBICACION,
            "plazo": data.PLAZO, "realizado_por": data.REALIZADO_POR,
            "fecha_realizacion": data.FECHA_REALIZACION,
            "observaciones": data.OBSERVACIONES,
            "nombre_dispositivo": data.nombre_dispositivo,
            "planta": data.PLANTA, "categoria_color": data.CATEGORIA_COLOR,
            "anio_creacion": data.ANIO_CREACION
        }
        usr = _obtener_usuario(request, usuario)
        _registrar_auditoria(id, usr, anterior, nuevo)
        return {"mensaje": "ACTUALIZADO"}
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════
# DELETE /PREVENTIVO/{id}
# ════════════════════════════════════════════════════════
@router.delete("/PREVENTIVO/{id}")
def eliminar_preventivo(id: int, request: Request):

    usuario = request.cookies.get("usuario")
    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT rol FROM public.usuarios WHERE usuario=%s", (usuario,))
                row = cursor.fetchone()
                if not row or row[0] != "ADMIN":
                    return {"error": "No tienes permiso para eliminar"}

                cursor.execute("DELETE FROM public.mantenimientos_preventivos WHERE id=%s", (id,))
                conn.commit()
            finally:
                cursor.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════
# PDF
# ════════════════════════════════════════════════════════
@router.post("/PREVENTIVO/PDF/{id}")
async def subir_pdf_preventivo(id: int, file: UploadFile = File(...)):
    try:
        contenido = await file.read()
        ruta = _pdf_path(id)
        ruta.write_bytes(contenido)

        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute('UPDATE public.mantenimientos_preventivos SET pdf=%s WHERE id=%s',
                               (str(ruta), id))
                conn.commit()
            finally:
                cursor.close()
        return {"mensaje": "PDF subido"}
    except Exception as e:
        return {"error": str(e)}


@router.get("/PREVENTIVO/PDF/{id}")
def obtener_pdf_preventivo(id: int):
    ruta = _pdf_path(id)
    if not ruta.exists():
        return {"error": "PDF no encontrado"}
    return Response(content=ruta.read_bytes(), media_type="application/pdf")


@router.delete("/PREVENTIVO/PDF/{id}")
def eliminar_pdf_preventivo(id: int):
    try:
        ruta = _pdf_path(id)
        if ruta.exists(): ruta.unlink()
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute('UPDATE public.mantenimientos_preventivos SET pdf=NULL WHERE id=%s', (id,))
                conn.commit()
            finally:
                cursor.close()
        return {"mensaje": "PDF eliminado"}
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════
# EXPORTAR
# ════════════════════════════════════════════════════════
@router.get("/PREVENTIVOS/EXPORTAR_TODO")
def exportar_preventivos_todo():

    with _db() as conn:
        df = pd.read_sql(
            'SELECT * FROM public.mantenimientos_preventivos ORDER BY id DESC',
            conn
        )

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        df.to_excel(writer, index=False, sheet_name="Preventivos")

        ws = writer.book.active

        for row in range(2, len(df)+2):

            color = str(df.iloc[row-2]["categoria_color"]).lower()

            fill = None

            if "verde" in color:
                fill = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")

            elif "amarillo" in color:
                fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

            elif "rojo" in color:
                fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")

            elif "gris" in color:
                fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

            elif "rosa" in color:
                fill = PatternFill(start_color="FFC5D3", end_color="FFC5D3", fill_type="solid")

            elif "azul" in color:
                fill = PatternFill(start_color="BEDFFB", end_color="BEDFFB", fill_type="solid")

            if fill:
                for col in range(1, len(df.columns)+1):
                    ws.cell(row=row, column=col).fill = fill

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=preventivos_todo.xlsx"}
    )


##################################EXPORTAR FILTRADO###################

@router.get("/PREVENTIVOS/EXPORTAR_FILTRADO")
def exportar_preventivos_filtrado(
    ID_EQUIPO: Optional[str] = Query(None),
    UBICACION: Optional[str] = Query(None),
    nombre_dispositivo: Optional[str] = Query(None),
    PLANTA: Optional[str] = Query(None),
    CATEGORIA_COLOR: Optional[str] = Query(None),
    OBSERVACIONES: Optional[str] = Query(None)
):

    where = "WHERE 1=1"
    params = []

    if ID_EQUIPO:
        where += ' AND id_equipo ILIKE %s'
        params.append(f"%{ID_EQUIPO}%")

    if UBICACION:
        where += ' AND ubicacion ILIKE %s'
        params.append(f"%{UBICACION}%")

    if nombre_dispositivo:
        where += ' AND nombre_dispositivo ILIKE %s'
        params.append(f"%{nombre_dispositivo}%")

    if PLANTA:
        where += ' AND planta ILIKE %s'
        params.append(f"%{PLANTA}%")

    if CATEGORIA_COLOR:
        where += ' AND categoria_color ILIKE %s'
        params.append(f"%{CATEGORIA_COLOR}%")

    if OBSERVACIONES:
        where += ' AND observaciones ILIKE %s'
        params.append(f"%{OBSERVACIONES}%")

    query = f"""
    SELECT
    id,
    id_equipo,
    ubicacion,
    nombre_dispositivo,
    planta,
    categoria_color,
    observaciones
    FROM public.mantenimientos_preventivos
    {where}
    ORDER BY id DESC
    """

    with _db() as conn:
        df = pd.read_sql(query, conn, params=params)

    # ───────── COLORES ─────────
    def aplicar_color(row):

        color = str(row["categoria_color"]).lower()

        if "verde" in color:
            return ['background-color: #bbf7d0']*len(row)

        if "amarillo" in color:
            return ['background-color: #fef9c3']*len(row)

        if "rojo" in color:
            return ['background-color: #fecaca']*len(row)

        if "gris" in color:
            return ['background-color: #e5e7eb']*len(row)

        if "rosa" in color:
            return ['background-color: #ffc5d3']*len(row)

        if "azul" in color:
            return ['background-color: #bedffb']*len(row)

        return ['']*len(row)

    styled = df.style.apply(aplicar_color, axis=1)

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        styled.to_excel(writer, index=False, sheet_name="Preventivos")

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=preventivos_filtrados.xlsx"
        }
    )

@router.get("/PREVENTIVOS/EXPORTAR_ANIO")
def exportar_preventivos_anio(anio: int = Query(...)):

    with _db() as conn:
        df = pd.read_sql("""
            SELECT *
            FROM public.mantenimientos_preventivos
            WHERE anio_creacion = %s
            ORDER BY id DESC
        """, conn, params=(anio,))

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        df.to_excel(writer, index=False, sheet_name="Preventivos")

        ws = writer.book.active

        for row in range(2, len(df)+2):

            color = str(df.iloc[row-2]["categoria_color"]).lower()

            fill = None

            if "verde" in color:
                fill = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")

            elif "amarillo" in color:
                fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

            elif "rojo" in color:
                fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")

            elif "gris" in color:
                fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

            elif "rosa" in color:
                fill = PatternFill(start_color="FFC5D3", end_color="FFC5D3", fill_type="solid")

            elif "azul" in color:
                fill = PatternFill(start_color="BEDFFB", end_color="BEDFFB", fill_type="solid")

            if fill:
                for col in range(1, len(df.columns)+1):
                    ws.cell(row=row, column=col).fill = fill

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=preventivos_{anio}.xlsx"
        }
    )


############################################
############ GENERAR QR POR UBICACION ######
############################################



@router.get("/QR_MESA_GENERAR/{ubicacion}")
def generar_qr_mesa(ubicacion: str):

    base = "http://172.24.104.166:8000"

    ubicacion_url = quote(ubicacion)
    url = f"{base}/preventivos/qr/{ubicacion_url}"

    qr = qrcode.make(url).convert("RGB")

    qr_width, qr_height = qr.size

    # espacio para texto
    altura_total = qr_height + 60
    img = Image.new("RGB", (qr_width, altura_total), "white")

    # pegar QR
    img.paste(qr, (0, 0))

    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("arial.ttf", 22)
    except:
        font = ImageFont.load_default()

    # centrar texto
    text_width = draw.textlength(ubicacion, font=font)
    x = (qr_width - text_width) / 2
    y = qr_height + 15

    draw.text((x, y), ubicacion, fill="black", font=font)

    # guardar en memoria
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")

    return Response(content=buffer.getvalue(), media_type="image/png")
#########################preventivo digital################
@router.post("/PREVENTIVO/GUARDAR_DIGITAL/{id}")
def guardar_preventivo_digital(id:int, data:dict):
    """
    Guarda el preventivo digital.
    - Guarda el JSON en el registro principal (id).
    - Actualiza fecha_realizacion=NOW() SOLO a los ids que vienen
      en data['ids_con_check'] (lista de ints).
      Si no viene esa clave, solo actualiza el registro principal
      (comportamiento anterior, por compatibilidad).
    """

    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                UPDATE public.mantenimientos_preventivos
                SET preventivo_digital = %s
                WHERE id = %s
                """, (json.dumps(data), id))

                ids_con_check = data.get("ids_con_check", [])

                print(f"[DEBUG GUARDAR] id principal: {id}")
                print(f"[DEBUG GUARDAR] ids_con_check raw: {ids_con_check}")
                print(f"[DEBUG GUARDAR] data keys: {list(data.keys())}")

                # Convertir a int de forma segura (JS puede enviar floats como 123.0)
                ids_validos = []
                for i in ids_con_check:
                    try:
                        ids_validos.append(int(float(i)))
                    except (TypeError, ValueError):
                        pass

                print(f"[DEBUG GUARDAR] ids_validos: {ids_validos}")

                if ids_validos:
                    placeholders = ",".join(["%s"] * len(ids_validos))
                    cursor.execute(f"""
                    UPDATE public.mantenimientos_preventivos
                    SET fecha_realizacion = NOW()
                    WHERE id IN ({placeholders})
                    """, ids_validos)
                    print(f"[DEBUG GUARDAR] UPDATE ejecutado para ids: {ids_validos}")
                else:
                    # Sin ids_con_check válidos → actualizar solo el registro principal
                    cursor.execute("""
                    UPDATE public.mantenimientos_preventivos
                    SET fecha_realizacion = NOW()
                    WHERE id = %s
                    """, (id,))
                    print(f"[DEBUG GUARDAR] UPDATE fallback solo id principal: {id}")

                conn.commit()
            finally:
                cursor.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}
@router.get("/PREVENTIVO/DIGITAL/{id}")
def obtener_preventivo_digital(id:int):

    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                SELECT preventivo_digital
                FROM public.mantenimientos_preventivos
                WHERE id=%s
                """, (id,))
                row = cursor.fetchone()
            finally:
                cursor.close()

        if not row or not row[0]:
            return {"existe": False}

        data = row[0]
        if isinstance(data, str):
            data = json.loads(data)

        return {"existe": True, "data": data}
    except Exception as e:
        return {"existe": False, "error": str(e)}
@router.delete("/PREVENTIVO/ELIMINAR_DIGITAL/{id}")
def eliminar_preventivo_digital(id:int):

    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                UPDATE public.mantenimientos_preventivos
                SET preventivo_digital = NULL
                WHERE id = %s
                """, (id,))
                conn.commit()
            finally:
                cursor.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


# ════════════════════════════════════════════════════════
# GET  /PREVENTIVO/VERIFICAR_USUARIO
# ════════════════════════════════════════════════════════
@router.get("/PREVENTIVO/VERIFICAR_USUARIO")
def verificar_usuario(usuario: str = Query(...)):
    try:
        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                SELECT nombre FROM public.usuarios
                WHERE usuario = %s AND activo = true
                """, (usuario.upper(),))
                row = cursor.fetchone()
            finally:
                cursor.close()
        if row:
            return {"existe": True, "nombre": row[0]}
        return {"existe": False}
    except Exception as e:
        return {"existe": False, "error": str(e)}


# ════════════════════════════════════════════════════════
# POST /PREVENTIVO/GUARDAR_PM/{id}
# Guarda el preventivo de una tarjeta individual y calcula
# automaticamente el plazo = fecha_realizacion + 6 meses
# ════════════════════════════════════════════════════════
@router.post("/PREVENTIVO/GUARDAR_PM/{id}")
def guardar_pm_individual(id: int, data: dict):
    try:
        usuario      = data.get("usuario", "SISTEMA")
        fecha_str    = data.get("fecha")
        checks       = data.get("checks", [])
        observaciones = data.get("observaciones", "")

        if not fecha_str:
            return {"ok": False, "error": "Fecha requerida"}

        fecha_realizacion = datetime.strptime(fecha_str, "%Y-%m-%d")
        proximo_pm        = fecha_realizacion + relativedelta(months=6)
        proximo_pm_str    = proximo_pm.strftime("%Y-%m-%d")

        preventivo_json = json.dumps({
            "usuario":       usuario,
            "fecha":         fecha_str,
            "proximo_pm":    proximo_pm_str,
            "checks":        checks,
            "observaciones": observaciones
        }, ensure_ascii=False)

        with _db() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("""
                UPDATE public.mantenimientos_preventivos
                SET fecha_realizacion  = %s,
                    plazo              = %s,
                    realizado_por      = %s,
                    observaciones      = %s,
                    preventivo_digital = %s
                WHERE id = %s
                """, (fecha_realizacion, proximo_pm_str, usuario.upper(), observaciones, preventivo_json, id))
                conn.commit()
            finally:
                cursor.close()

        return {"ok": True, "proximo_pm": proximo_pm_str}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.get("/preventivos/qr/{ubicacion}", response_class=HTMLResponse)
def ver_qr_preventivo(ubicacion: str):

    with _db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
            SELECT id, id_equipo, nombre_dispositivo, planta,
                   categoria_color, fecha_realizacion, plazo, observaciones,
                   CASE WHEN preventivo_digital IS NOT NULL THEN true ELSE false END AS tiene_pm
            FROM public.mantenimientos_preventivos
            WHERE ubicacion=%s
            ORDER BY nombre_dispositivo
            """, (ubicacion,))
            rows = cursor.fetchall()
        finally:
            cursor.close()

    def color_badge(cat):
        c = (cat or "").lower()
        if "verde"    in c: return "#10B981","#052e16","Verde"
        if "amarillo" in c: return "#F59E0B","#1c1400","Amarillo"
        if "rojo"     in c: return "#EF4444","#1f0000","Rojo"
        if "gris"     in c: return "#94A3B8","#0f172a","Gris"
        if "rosa"     in c: return "#F472B6","#1f0011","Rosa"
        if "azul"     in c: return "#3B82F6","#001233","Azul"
        return "#64748B","#0f172a", cat or "—"

    def disp_icon(disp):
        d = (disp or "").upper()
        if "COMPUTADORA" in d or "CPU" in d: return "🖥️"
        if "PORTATIL" in d or "LAPTOP" in d: return "💻"
        if "IMPRESORA" in d:                 return "🖨️"
        if "UPS" in d:                       return "🔋"
        return "🔧"

    def actividades_dispositivo(disp):
        d = (disp or "").upper()
        if "COMPUTADORA" in d or "CPU" in d:
            return [
                "Sopletear el gabinete",
                "Limpieza de contactos de memoria RAM",
                "Sopletear fuente de poder y ventiladores",
                "Limpieza del gabinete",
                "Limpieza del monitor o pantalla",
                "Limpieza y sopleteado del teclado y mouse",
                "Sopleteado de ventiladores y ranuras de enfriamiento",
                "Limpieza exterior del lector óptico",
                "Limpieza del cableado",
                "Actualizaciones del sistema operativo",
                "Actualizaciones de Office",
                "Eliminación de archivos temporales y vaciar reciclaje",
                "Revisión del estado del antivirus y escaneo",
                "Desfragmentar las unidades de disco duro",
                "Conectar todos los periféricos correspondientes",
                "Verificar cables y conectores sin daños",
                "Encender el equipo y verificar funcionamiento",
                "Verificar que los periféricos funcionen correctamente",
                "Verificación vida de la pila del BIOS",
            ]
        if "PORTATIL" in d or "LAPTOP" in d:
            return [
                "Sopletear el gabinete / chasis",
                "Limpieza de contactos de memoria RAM",
                "Sopletear fuente de poder y ventiladores",
                "Limpieza del monitor o pantalla",
                "Limpieza y sopleteado del teclado y touchpad",
                "Sopleteado de ventiladores y ranuras de enfriamiento",
                "Limpieza del cableado",
                "Actualizaciones del sistema operativo",
                "Actualizaciones de Office",
                "Eliminación de archivos temporales y vaciar reciclaje",
                "Revisión del estado del antivirus y escaneo",
                "Desfragmentar las unidades de disco duro",
                "Conectar todos los periféricos correspondientes",
                "Verificar cables y conectores sin daños",
                "Encender el equipo y verificar funcionamiento",
                "Verificar que los periféricos funcionen correctamente",
            ]
        if "IMPRESORA" in d:
            return [
                "Sopletear la impresora térmica",
                "Limpieza de rodillos (no usar alcohol)",
                "Limpieza del cabezal de la impresora térmica",
                "Limpieza exterior de la impresora",
                "Limpieza del cableado",
                "Rutear cables / anclar eliminador de impresora",
                "Conectar todos los periféricos correspondientes",
                "Verificar cables y conectores sin daños",
                "Verificar que los periféricos funcionen correctamente",
            ]
        if "UPS" in d:
            return [
                "Limpieza y verificación del UPS",
                "Limpieza del cableado",
                "Conectar todos los periféricos correspondientes",
                "Verificar cables y conectores sin daños",
                "Verificación vida de la pila del UPS",
                "Inspección y funcionamiento del UPS",
                "Verificar que solo equipo IT esté conectado al UPS",
            ]
        return ["Inspección general", "Limpieza exterior", "Verificación de funcionamiento"]

    cards_html = ""
    for r in rows:
        id_registro = r[0]
        id_equipo   = r[1] or ""
        dispositivo = r[2] or ""
        planta      = r[3] or ""
        color_cat   = r[4] or ""
        fecha       = r[5]
        plazo       = r[6]

        badge_color, badge_bg, badge_label = color_badge(color_cat)
        icon      = disp_icon(dispositivo)
        fecha_str = str(fecha)[:10] if fecha else "Sin registro"
        plazo_str = str(plazo)[:10] if plazo else "No definido"
        dot_class = "dot-ok" if fecha else "dot-warn"
        dot_label = f"Último PM: {fecha_str}" if fecha else "Sin mantenimiento registrado"

        tiene_pm  = r[8] if len(r) > 8 else False
        acts = actividades_dispositivo(dispositivo)
        acts_html = "".join([
            f'<label class="act-item"><input type="checkbox"><span class="act-check"></span><span class="act-text">{a}</span></label>'
            for a in acts
        ])
        if tiene_pm:
            btn_pm = (
                f'<button class="pm-btn btn btn-cyan" onclick="verPM({id_registro})" style="display:none">👁 Ver PM</button>\n    '
                f'<button class="pm-btn btn btn-amber" onclick="abrirEditarPM({id_registro})" style="display:none">✏️ Editar PM</button>\n    '
                f'<button class="pm-btn btn btn-danger" onclick="eliminarPreventivo({id_registro})" style="display:none">🗑 Eliminar</button>'
            )
        else:
            btn_pm = f'<button class="pm-btn btn btn-purple" onclick="abrirForm({id_registro})" style="display:none">📋 Hacer Preventivo</button>'

        cards_html += f"""
<div class="card">
  <input type="hidden" id="ubicacion_{id_registro}" value="{ubicacion}">
  <div class="card-top">
    <div class="dev-icon">{icon}</div>
    <div class="dev-name"><h3>{dispositivo or "Dispositivo"}</h3><span>{id_equipo}</span></div>
    <span class="color-badge" style="background:{badge_bg};color:{badge_color};border:1px solid {badge_color}40">{badge_label}</span>
  </div>
  <div class="card-body">
    <div class="info-row">
      <div class="info-item"><label>ID Equipo</label><input id="equipo_{id_registro}" value="{id_equipo}" disabled></div>
      <div class="info-item"><label>Dispositivo</label><input id="disp_{id_registro}" value="{dispositivo}" disabled></div>
      <div class="info-item"><label>Planta</label><input id="planta_{id_registro}" value="{planta}" disabled></div>
      <div class="info-item"><label>Color</label><input id="color_{id_registro}" value="{color_cat}" disabled></div>
    </div>
    <div class="status-row">
      <span class="status-dot {dot_class}"></span>
      <span>{dot_label}</span>
      <span style="margin-left:auto;font-family:'DM Mono',monospace;font-size:10px;color:#475569">Plazo: {plazo_str}</span>
    </div>
    <div id="obswrap_{id_registro}" style="display:none;margin-top:10px">
      <div class="obs-label">Observaciones</div>
      <textarea class="obs-edit-field" id="obs_{id_registro}" disabled>{r[7] or ""}</textarea>
    </div>
    <div class="mini-form" id="form_{id_registro}" style="display:none">
      <div class="form-sep" style="margin-top:12px">📋 Actividades</div>
      <div class="acts-list">{acts_html}</div>
      <div class="form-sep" style="margin-top:10px">📅 Fecha</div>
      <input type="date" class="date-input" id="fecha_{id_registro}">
      <div class="form-sep" style="margin-top:8px">📝 Observaciones</div>
      <textarea class="date-input" style="min-height:52px;resize:vertical;font-family:'DM Sans',sans-serif;" id="obs_pm_{id_registro}" placeholder="Observaciones..."></textarea>
      <div class="form-actions" style="margin-top:8px">
        <button class="btn btn-ghost" onclick="cancelarForm({id_registro})">✕ Cancelar</button>
        <button class="btn btn-success" onclick="guardarPreventivo({id_registro})">💾 Guardar</button>
      </div>
    </div>
    <div class="mini-form" id="ver_{id_registro}" style="display:none">
      <div class="form-sep" style="margin-top:12px;color:#06B6D4">👁 Preventivo registrado</div>
      <div class="acts-list" id="ver_acts_{id_registro}" style="max-height:180px;overflow-y:auto"></div>
      <div style="margin-top:10px;font-size:11px;color:var(--muted2);display:flex;flex-wrap:wrap;gap:8px">
        <span>👤 <b id="ver_usuario_{id_registro}"></b></span>
        <span>📅 <b id="ver_fecha_{id_registro}"></b></span>
        <span>⏭ <b id="ver_proximo_{id_registro}"></b></span>
      </div>
      <div style="margin-top:6px;font-size:11px;color:var(--muted2)" id="ver_obs_{id_registro}"></div>
      <div class="form-actions" style="margin-top:8px">
        <button class="btn btn-ghost" onclick="cerrarVer({id_registro})">✕ Cerrar</button>
      </div>
    </div>
    <div class="mini-form" id="edit_pm_{id_registro}" style="display:none">
      <div class="form-sep" style="margin-top:12px;color:var(--amber)">✏️ Editar Preventivo</div>
      <div class="acts-list" id="edit_acts_{id_registro}">{acts_html}</div>
      <div class="form-sep" style="margin-top:10px">📅 Fecha</div>
      <input type="date" class="date-input" id="edit_fecha_{id_registro}">
      <div class="form-sep" style="margin-top:8px">📝 Observaciones</div>
      <textarea class="date-input" style="min-height:52px;resize:vertical;font-family:'DM Sans',sans-serif;" id="edit_obs_pm_{id_registro}" placeholder="Observaciones..."></textarea>
      <div class="form-actions" style="margin-top:8px">
        <button class="btn btn-ghost" onclick="cerrarEditarPM({id_registro})">✕ Cancelar</button>
        <button class="btn btn-amber" onclick="guardarEditarPM({id_registro})">💾 Guardar Cambios</button>
      </div>
    </div>
  </div>
  <div class="card-actions">
    <button class="btn btn-blue" onclick="abrirEditar({id_registro})">✏️ Editar</button>
    <button class="btn btn-green" onclick="guardarCambios({id_registro})">💾 Guardar</button>
    <button class="btn btn-ghost" onclick="cancelarEditar({id_registro})">↩ Cancelar</button>
    <button class="btn btn-ghost" onclick="window.close()">✕ Salir</button>
    {btn_pm}
  </div>
</div>"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PM — {ubicacion}</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{{--bg:#0B0F1A;--surface:#111827;--surface2:#1a2235;--border:rgba(255,255,255,0.07);--border2:rgba(255,255,255,0.12);--accent:#3B82F6;--text:#F1F5F9;--muted:#64748B;--muted2:#94A3B8;--green:#10B981;--red:#EF4444;--amber:#F59E0B;--radius:14px;}}
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;padding-bottom:40px;}}
.top-bar{{background:linear-gradient(135deg,#0f1e35,#0B0F1A);border-bottom:1px solid var(--border2);padding:14px 20px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:100;}}
.top-icon{{width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,#1D4ED8,#3B82F6);display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0;box-shadow:0 0 16px rgba(59,130,246,.4);}}
.top-title{{flex:1;}}
.top-title h1{{font-size:15px;font-weight:700;}}
.top-title p{{font-size:11px;color:var(--muted2);margin-top:2px;font-family:'DM Mono',monospace;}}
.user-chip{{display:none;align-items:center;gap:8px;padding:7px 14px;border-radius:999px;background:rgba(16,185,129,.15);border:1px solid rgba(16,185,129,.3);font-size:12px;font-weight:600;color:#6ee7b7;}}
.btn{{display:inline-flex;align-items:center;gap:6px;padding:9px 16px;border:none;border-radius:8px;font-family:'DM Sans',sans-serif;font-size:13px;font-weight:600;cursor:pointer;transition:all .2s;}}
.btn-primary{{background:var(--accent);color:white;box-shadow:0 4px 12px rgba(59,130,246,.35);}}
.btn-primary:hover{{background:#2563EB;transform:translateY(-1px);}}
.btn-success{{background:var(--green);color:white;}}
.btn-success:hover{{background:#059669;transform:translateY(-1px);}}
.btn-ghost{{background:var(--surface2);color:var(--muted2);border:1px solid var(--border2);}}
.btn-ghost:hover{{color:var(--text);border-color:var(--accent);}}
.btn-blue{{background:#3B82F6;color:white;box-shadow:0 3px 10px rgba(59,130,246,.3);}}
.btn-blue:hover{{background:#2563EB;transform:translateY(-1px);}}
.btn-green{{background:#10B981;color:white;box-shadow:0 3px 10px rgba(16,185,129,.3);}}
.btn-green:hover{{background:#059669;transform:translateY(-1px);}}
.btn-amber{{background:var(--amber);color:#1c1400;box-shadow:0 3px 10px rgba(245,158,11,.3);}}
.btn-amber:hover{{background:#D97706;transform:translateY(-1px);}}
.btn-danger{{background:var(--red);color:white;box-shadow:0 3px 10px rgba(239,68,68,.3);}}
.btn-danger:hover{{background:#DC2626;transform:translateY(-1px);}}
.btn-cyan{{background:#06B6D4;color:#001a1f;box-shadow:0 3px 10px rgba(6,182,212,.3);}}
.btn-cyan:hover{{background:#0891B2;transform:translateY(-1px);}}
.btn-purple{{background:#8B5CF6;color:white;box-shadow:0 3px 10px rgba(139,92,246,.3);}}
.btn-purple:hover{{background:#7C3AED;transform:translateY(-1px);}}
.btn-login{{background:linear-gradient(135deg,#1D4ED8,#3B82F6);color:white;box-shadow:0 4px 14px rgba(59,130,246,.4);}}
.btn-login:hover{{transform:translateY(-1px);}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:14px;padding:16px 20px;}}
.card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden;transition:border-color .2s;animation:fadeUp .35s ease both;}}
@keyframes fadeUp{{from{{opacity:0;transform:translateY(12px)}}to{{opacity:1;transform:translateY(0)}}}}
.card-top{{display:flex;align-items:center;gap:12px;padding:14px 16px;border-bottom:1px solid var(--border);background:var(--surface2);}}
.dev-icon{{width:38px;height:38px;border-radius:9px;background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.2);display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0;}}
.dev-name{{flex:1;}}
.dev-name h3{{font-size:13px;font-weight:700;}}
.dev-name span{{font-size:11px;color:var(--muted2);font-family:'DM Mono',monospace;}}
.color-badge{{padding:4px 10px;border-radius:999px;font-size:10px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;}}
.card-body{{padding:14px 16px;}}
.info-row{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px;}}
.info-item label{{font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--muted);display:block;margin-bottom:4px;}}
.info-item input{{width:100%;background:var(--surface2);border:1px solid var(--border2);border-radius:6px;padding:7px 9px;font-size:12px;font-family:'DM Mono',monospace;color:var(--text);opacity:.6;}}
.status-row{{display:flex;align-items:center;gap:8px;padding:9px 12px;background:rgba(255,255,255,.03);border:1px solid var(--border);border-radius:8px;margin-bottom:12px;font-size:11px;color:var(--muted2);}}
.status-dot{{width:7px;height:7px;border-radius:50%;flex-shrink:0;}}
.dot-ok{{background:var(--green);box-shadow:0 0 6px var(--green);}}
.dot-warn{{background:var(--amber);box-shadow:0 0 6px var(--amber);}}
.pm-btn{{font-size:12px !important;padding:8px 14px !important;}}
.mini-form{{border-top:1px solid var(--border);padding-top:12px;margin-top:4px;}}
.form-sep{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--accent);margin-bottom:8px;}}
.acts-list{{display:flex;flex-direction:column;gap:3px;margin-bottom:4px;max-height:220px;overflow-y:auto;}}
.act-item{{display:flex;align-items:flex-start;gap:8px;padding:5px 8px;border-radius:6px;cursor:pointer;transition:background .15s;font-size:12px;color:var(--muted2);}}
.act-item:hover{{background:rgba(59,130,246,.08);color:var(--text);}}
.act-item input[type=checkbox]{{display:none;}}
.act-check{{width:16px;height:16px;border-radius:4px;border:1.5px solid rgba(59,130,246,.3);background:var(--surface2);flex-shrink:0;margin-top:2px;display:flex;align-items:center;justify-content:center;font-size:10px;color:transparent;transition:all .15s;}}
.act-item input:checked ~ .act-check{{background:var(--accent);border-color:var(--accent);color:white;}}
.act-item input:checked ~ .act-check::after{{content:"✓";}}
.act-item input:checked ~ .act-text{{color:var(--text);}}
.date-input{{width:100%;background:var(--surface2);border:1px solid var(--border2);border-radius:6px;padding:8px 10px;font-size:12px;font-family:'DM Mono',monospace;color:var(--text);color-scheme:dark;margin-bottom:8px;}}
.date-input:focus{{outline:none;border-color:var(--accent);}}
.obs-field{{width:100%;background:var(--surface2);border:1px solid var(--border2);border-radius:6px;padding:8px 10px;font-size:12px;color:var(--text);resize:vertical;min-height:52px;font-family:'DM Sans',sans-serif;margin-bottom:10px;}}
.obs-field:focus{{outline:none;border-color:var(--accent);}}
.form-actions{{display:flex;gap:8px;justify-content:flex-end;}}
.form-actions .btn{{font-size:12px;padding:8px 14px;}}
.modal{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);backdrop-filter:blur(4px);justify-content:center;align-items:center;z-index:9999;}}
.modal.show{{display:flex;}}
.modal-box{{background:var(--surface);border:1px solid var(--border2);border-radius:16px;padding:28px;width:min(360px,95vw);box-shadow:0 30px 80px rgba(0,0,0,.6);animation:fadeUp .2s ease;}}
.modal-box h3{{font-size:16px;font-weight:700;margin-bottom:6px;}}
.modal-box p{{font-size:12px;color:var(--muted2);margin-bottom:20px;}}
.modal-field{{margin-bottom:16px;}}
.modal-field label{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--muted);display:block;margin-bottom:5px;}}
.modal-field input{{width:100%;background:var(--surface2);border:1px solid var(--border2);border-radius:7px;padding:10px 12px;font-size:14px;font-family:'DM Mono',monospace;color:var(--text);}}
.modal-field input:focus{{outline:none;border-color:var(--accent);box-shadow:0 0 0 3px rgba(59,130,246,.15);}}
.modal-footer{{display:flex;gap:8px;justify-content:flex-end;}}
.toast{{position:fixed;bottom:24px;right:24px;padding:12px 20px;border-radius:10px;font-size:13px;font-weight:600;z-index:99999;animation:fadeUp .3s ease;pointer-events:none;}}
.toast-ok{{background:#052e16;border:1px solid #10B981;color:#6ee7b7;}}
.toast-err{{background:#1f0000;border:1px solid #EF4444;color:#fca5a5;}}
</style>
</head>
<body>
<div class="top-bar">
  <div class="top-icon">🔧</div>
  <div class="top-title"><h1>Mantenimiento Preventivo</h1><p>📍 {ubicacion}</p></div>
  <div class="user-chip" id="userChip">👤 <span id="userNombre"></span></div>
  <button class="btn btn-login" id="btnLogin" onclick="abrirLogin()">🔑 Iniciar Sesión</button>
</div>

<div class="grid">
{cards_html}
</div>

<!-- Modal Login -->
<div class="modal" id="modalLogin">
  <div class="modal-box">
    <h3>🔑 Iniciar Sesión</h3>
    <p>Ingresa tu usuario para habilitar el registro de preventivos</p>
    <div class="modal-field">
      <label>Usuario</label>
      <input id="inputUsuario" type="text" placeholder="Ej: DOMINGUEZG" autocomplete="off"
             onkeydown="if(event.key==='Enter') confirmarLogin()">
    </div>
    <div class="modal-footer">
      <button class="btn btn-ghost" onclick="cerrarLogin()">Cancelar</button>
      <button class="btn btn-primary" onclick="confirmarLogin()">Entrar →</button>
    </div>
  </div>
</div>

<script>
let usuarioActual  = null;
let nombreActual   = null;
let usuarioTarjeta = {{}};

function abrirLogin() {{
  document.getElementById('modalLogin').classList.add('show');
  setTimeout(() => document.getElementById('inputUsuario').focus(), 100);
}}
function cerrarLogin() {{
  document.getElementById('modalLogin').classList.remove('show');
  document.getElementById('inputUsuario').value = '';
}}
async function confirmarLogin() {{
  const usr = document.getElementById('inputUsuario').value.trim().toUpperCase();
  if (!usr) {{ toast('Ingresa tu usuario', false); return; }}
  const res  = await fetch('/PREVENTIVO/VERIFICAR_USUARIO?usuario=' + encodeURIComponent(usr));
  const data = await res.json();
  if (!data.existe) {{ toast('❌ Usuario no encontrado', false); return; }}
  usuarioActual = usr;
  nombreActual  = data.nombre || usr;
  document.getElementById('userNombre').textContent = nombreActual;
  document.getElementById('userChip').style.display = 'flex';
  document.getElementById('btnLogin').style.display  = 'none';
  document.querySelectorAll('.pm-btn').forEach(b => b.style.display = 'inline-flex');
  cerrarLogin();
  toast('✅ Sesión iniciada — ' + nombreActual, true);
}}

// ── HACER PREVENTIVO ──
function abrirForm(id) {{
  if (!usuarioActual) {{ abrirLogin(); return; }}
  document.getElementById('form_' + id).style.display = 'block';
  document.getElementById('fecha_' + id).value = new Date().toISOString().split('T')[0];
}}
function cancelarForm(id) {{
  document.getElementById('form_' + id).style.display = 'none';
  document.querySelectorAll('#form_' + id + ' input[type=checkbox]').forEach(cb => cb.checked = false);
  const o = document.getElementById('obs_pm_' + id); if (o) o.value = '';
}}
async function guardarPreventivo(id) {{
  const fecha = document.getElementById('fecha_' + id).value;
  if (!fecha) {{ toast('Selecciona la fecha', false); return; }}
  const cbs = document.querySelectorAll('#form_' + id + ' input[type=checkbox]');
  const checks = []; cbs.forEach((cb,i) => {{ if (cb.checked) checks.push(i); }});
  if (!checks.length) {{ toast('Marca al menos una actividad', false); return; }}
  const obs = document.getElementById('obs_pm_' + id)?.value || '';
  const btn = document.querySelector('#form_' + id + ' .btn-success');
  btn.disabled = true; btn.textContent = 'Guardando...';
  const res  = await fetch('/PREVENTIVO/GUARDAR_PM/' + id, {{
    method:'POST', headers:{{'Content-Type':'application/json'}},
    body: JSON.stringify({{ usuario: usuarioActual, fecha, checks, observaciones: obs }})
  }});
  const data = await res.json();
  if (data.ok) {{
    toast('✅ Guardado. Próximo PM: ' + data.proximo_pm, true);
    const card = document.getElementById('form_' + id).closest('.card');
    card.querySelector('.status-dot').className = 'status-dot dot-ok';
    card.querySelector('.status-row span:nth-child(2)').textContent = 'Último PM: ' + fecha;
    card.querySelector('.status-row span:last-child').textContent   = 'Plazo: ' + data.proximo_pm;
    cancelarForm(id);
  }} else {{
    btn.disabled = false; btn.textContent = '💾 Guardar';
    toast('❌ Error: ' + (data.error||'desconocido'), false);
  }}
}}

// ── VER PM ──
async function verPM(id) {{
  const res  = await fetch('/PREVENTIVO/DIGITAL/' + id);
  const data = await res.json();
  if (!data.existe) {{ toast('No hay preventivo guardado', false); return; }}
  const pm = data.data;
  const actsEl  = document.getElementById('ver_acts_' + id);
  const allActs = Array.from(document.querySelectorAll('#edit_acts_' + id + ' .act-text')).map(e => e.textContent);
  actsEl.innerHTML = '';
  allActs.forEach((act, i) => {{
    const marcado = pm.checks && pm.checks.includes(i);
    actsEl.innerHTML += `<div class="act-item" style="opacity:${{marcado?1:.35}}">
      <span class="act-check" style="${{marcado?'background:var(--accent);border-color:var(--accent);color:white':''}}">
        ${{marcado?'✓':''}}
      </span><span class="act-text">${{act}}</span></div>`;
  }});
  document.getElementById('ver_usuario_' + id).textContent = pm.usuario    || '—';
  document.getElementById('ver_fecha_'   + id).textContent = pm.fecha      || '—';
  document.getElementById('ver_proximo_' + id).textContent = pm.proximo_pm || '—';
  const o = document.getElementById('ver_obs_' + id);
  if (o) o.textContent = pm.observaciones ? '📝 ' + pm.observaciones : '';
  document.getElementById('ver_' + id).style.display = 'block';
}}
function cerrarVer(id) {{ document.getElementById('ver_' + id).style.display = 'none'; }}

// ── EDITAR PM ──
async function abrirEditarPM(id) {{
  if (!usuarioActual) {{ abrirLogin(); return; }}
  const res  = await fetch('/PREVENTIVO/DIGITAL/' + id);
  const data = await res.json();
  const cbs  = document.querySelectorAll('#edit_acts_' + id + ' input[type=checkbox]');
  cbs.forEach(cb => cb.checked = false);
  if (data.existe && data.data.checks) data.data.checks.forEach(i => {{ if(cbs[i]) cbs[i].checked=true; }});
  const f = document.getElementById('edit_fecha_'  + id); if(f) f.value = data.existe?(data.data.fecha||''):'';
  const o = document.getElementById('edit_obs_pm_' + id); if(o) o.value = data.existe?(data.data.observaciones||''):'';
  document.getElementById('edit_pm_' + id).style.display = 'block';
}}
function cerrarEditarPM(id) {{ document.getElementById('edit_pm_' + id).style.display = 'none'; }}
async function guardarEditarPM(id) {{
  const fecha = document.getElementById('edit_fecha_' + id).value;
  if (!fecha) {{ toast('Selecciona la fecha', false); return; }}
  const cbs = document.querySelectorAll('#edit_acts_' + id + ' input[type=checkbox]');
  const checks = []; cbs.forEach((cb,i) => {{ if(cb.checked) checks.push(i); }});
  if (!checks.length) {{ toast('Marca al menos una actividad', false); return; }}
  const obs = document.getElementById('edit_obs_pm_' + id).value;
  const btn = document.querySelector('#edit_pm_' + id + ' .btn-amber');
  btn.disabled = true; btn.textContent = 'Guardando...';
  const res  = await fetch('/PREVENTIVO/GUARDAR_PM/' + id, {{
    method:'POST', headers:{{'Content-Type':'application/json'}},
    body: JSON.stringify({{ usuario: usuarioActual, fecha, checks, observaciones: obs }})
  }});
  const data = await res.json();
  if (data.ok) {{
    toast('✅ Actualizado. Próximo PM: ' + data.proximo_pm, true);
    const card = document.getElementById('edit_pm_' + id).closest('.card');
    card.querySelector('.status-dot').className = 'status-dot dot-ok';
    card.querySelector('.status-row span:nth-child(2)').textContent = 'Último PM: ' + fecha;
    card.querySelector('.status-row span:last-child').textContent   = 'Plazo: ' + data.proximo_pm;
    cerrarEditarPM(id);
  }} else {{
    btn.disabled = false; btn.textContent = '💾 Guardar Cambios';
    toast('❌ Error: ' + (data.error||'desconocido'), false);
  }}
}}

// ── ELIMINAR PM ──
async function eliminarPreventivo(id) {{
  if (!confirm('¿Eliminar el preventivo guardado?')) return;
  const res  = await fetch('/PREVENTIVO/ELIMINAR_DIGITAL/' + id, {{method:'DELETE'}});
  const data = await res.json();
  if (data.ok) {{
    toast('✅ Preventivo eliminado', true);
    const card = document.getElementById('ver_' + id).closest('.card');
    // Cerrar paneles
    document.getElementById('ver_'    + id).style.display = 'none';
    document.getElementById('edit_pm_'+ id).style.display = 'none';
    // Restaurar status a sin mantenimiento
    card.querySelector('.status-dot').className = 'status-dot dot-warn';
    card.querySelector('.status-row span:nth-child(2)').textContent = 'Sin mantenimiento registrado';
    card.querySelector('.status-row span:last-child').textContent   = 'Plazo: No definido';
    // Cambiar botones PM
    card.querySelectorAll('.btn-cyan,.btn-danger').forEach(b => b.style.display='none');
    card.querySelectorAll('.pm-btn.btn-amber').forEach(b => b.style.display='none');
    card.querySelectorAll('.btn-purple').forEach(b => b.style.display='inline-flex');
  }} else {{
    toast('❌ Error al eliminar', false);
  }}
}}

// ── EDITAR TARJETA ──
function abrirEditar(id) {{
  if (!usuarioActual) {{ abrirLogin(); return; }}
  usuarioTarjeta[id] = usuarioActual;
  ['equipo_','disp_','planta_','color_','obs_'].forEach(p => document.getElementById(p+id).disabled=false);
  document.getElementById('obswrap_' + id).style.display = 'block';
}}
function cancelarEditar(id) {{
  ['equipo_','disp_','planta_','color_','obs_'].forEach(p => document.getElementById(p+id).disabled=true);
  document.getElementById('obswrap_' + id).style.display = 'none';
}}
async function guardarCambios(id) {{
  const datos = {{
    ID_EQUIPO:          document.getElementById('equipo_'    + id).value,
    UBICACION:          document.getElementById('ubicacion_' + id).value,
    nombre_dispositivo: document.getElementById('disp_'      + id).value,
    PLANTA:             document.getElementById('planta_'    + id).value,
    CATEGORIA_COLOR:    document.getElementById('color_'     + id).value,
    OBSERVACIONES:      document.getElementById('obs_'       + id).value
  }};
  const usuario = usuarioTarjeta[id] || usuarioActual || 'SISTEMA';
  const res  = await fetch('/PREVENTIVO/' + id + '?usuario=' + encodeURIComponent(usuario), {{
    method:'PUT', headers:{{'Content-Type':'application/json','X-Usuario':usuario}},
    body: JSON.stringify(datos)
  }});
  const data = await res.json();
  if (data.mensaje) {{ toast('✅ Cambios guardados', true); cancelarEditar(id); }}
  else {{ toast('❌ Error al guardar', false); }}
}}

function toast(msg, ok) {{
  const t = document.createElement('div');
  t.className = 'toast ' + (ok ? 'toast-ok' : 'toast-err');
  t.textContent = msg;
  document.body.appendChild(t);
  setTimeout(() => t.remove(), 3000);
}}
</script>
</body>
</html>"""

    return html